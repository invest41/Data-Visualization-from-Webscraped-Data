import Webscraping_Tool, Make_Translation
import openpyxl as oex

#Open a Workbook
nb = oex.Workbook()
nb.save('web_scrape.xlsx')

#Access on Sheet
wb = oex.load_workbook('web_scrape.xlsx')
sheet = wb['Sheet']

#Work on Sheet

#Assign Cell Value
sheet.cell(1,1).value = 'Word'
sheet.cell(1,2).value = 'Frequency'

rn = 2
for line in open('clean_work.txt'):
	line = line.split()
	cn = 1
	for i in line:
		sheet.cell(rn,cn).value = i
		cn += 1
	rn += 1

#Save Workbook
wb.save('web_scrape.xlsx')
